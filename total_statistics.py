import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

ROOT_DIR = Path("source")

UPPER_ORDER = ["1-1", "2-1", "3-1", "4-1"]
LOWER_ORDER = ["1-2", "2-2", "3-2", "4-2"]


# ✅ 1) 파일당 1번만 읽는 버전
def extract_stats_fast(xlsx_path):
    df = pd.read_excel(xlsx_path, header=None, nrows=2)  # 처음 2줄만 읽음
    first_row = df.iloc[0, :].dropna().tolist()   # 통계 항목 row
    second_row = df.iloc[1, :len(first_row)].tolist()  # 값 row
    return first_row, second_row


# ✅ 2) stats_cache 기반 테이블 생성 (read_excel 호출 없음)
def build_wide_table(order_list, stats_cache):
    first_key = next((k for k in order_list if k in stats_cache), None)

    if first_key:
        first_row, _ = stats_cache[first_key]
        num_stats = len(first_row)
    else:
        num_stats = 14  # fallback
        print(f"⚠️ 파일 없음 → 기본 14개 통계로 처리")

    # Row 1 : 통계 항목명
    first_data_row = ["종속변수"]
    for stat_idx in range(num_stats):
        for key in order_list:
            if key in stats_cache:
                first_row = stats_cache[key][0]
                first_data_row.append(first_row[stat_idx])
            else:
                first_data_row.append("")

    # Row 2 : 1-1, 2-1, 3-1, 4-1 (조건 반복)
    condition_row = [""] + order_list * num_stats

    # Row 3 : 실제 값
    second_data_row = [""]
    for stat_idx in range(num_stats):
        for key in order_list:
            if key in stats_cache:
                second_row = stats_cache[key][1]
                second_data_row.append(second_row[stat_idx])
            else:
                second_data_row.append("")

    df = pd.DataFrame([first_data_row, condition_row, second_data_row])
    return df, num_stats


# ✅ 3) 병합 함수 (그대로 유지)
def apply_merges(worksheet, start_row, num_stats):
    for i in range(num_stats):
        start_col = 2 + i * 4
        end_col = start_col + 3
        worksheet.merge_cells(start_row=start_row, start_column=start_col,
                              end_row=start_row, end_column=end_col)


# ✅ 4) 피험자별 처리
def process_subject(subject_path):
    subject_name = subject_path.name
    xlsx_files = {f.parent.name: f for f in subject_path.rglob("_speed_time.xlsx")}
    print(f"\n📌 처리 중: {subject_name} ({len(xlsx_files)}개 파일 감지)")

    # ✅ 파일당 1회만 읽는 캐시
    stats_cache = {key: extract_stats_fast(path) for key, path in xlsx_files.items()}

    print("  ➜ 상단 테이블 생성")
    upper_df, num_stats = build_wide_table(UPPER_ORDER, stats_cache)

    print("  ➜ 하단 테이블 생성")
    lower_df, _ = build_wide_table(LOWER_ORDER, stats_cache)

    output_path = subject_path / "total_speed_statistics.xlsx"

    # ✅ 엑셀 저장
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        upper_df.to_excel(writer, sheet_name="통계", index=False, header=False, startrow=0)
        lower_df.to_excel(writer, sheet_name="통계", index=False, header=False,
                          startrow=len(upper_df) + 2)

    # ✅ 병합 처리
    wb = load_workbook(output_path)
    ws = wb["통계"]

    apply_merges(ws, start_row=1, num_stats=num_stats)
    apply_merges(ws, start_row=len(upper_df) + 3, num_stats=num_stats)

    wb.save(output_path)
    print(f"✅ 완료: {output_path}")


# ✅ 5) 실행부
for subject_dir in ROOT_DIR.iterdir():
    if subject_dir.is_dir():
        process_subject(subject_dir)

print("\n🎉 전체 total_speed_statistics.xlsx 생성 완료")
